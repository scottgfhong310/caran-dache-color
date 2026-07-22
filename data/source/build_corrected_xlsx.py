#!/usr/bin/env python3
"""
build_corrected_xlsx.py — produce a corrected reference workbook (v1.1.0-corrected) from upstream v1.1.0.

Applies the SUP (Supracolor) & NC2 (Neocolor II) re-sampled hexes (resampled_hex.json,
from the official chart PDFs — see DESIGN.md §4.1) consistently across every sheet that
carries a hex, and recomputes all colour-derived columns so the workbook stays internally
consistent:
  · Series_Color_Index : css_hex_approx, rgb, hsl, lab, contrast(black/white/best),
                         relative_luminance, recommended_foreground, wcag_aa_normal_text
  · Cross_Series_Map   : the SUP / NC2 columns
  · Swatch_Reference   : css_hex_approx, recommended_foreground_hex (SUP/NC2 rows)
  · Color_Master       : for every canonical code that includes SUP or NC2 —
                         avg hex/rgb/lab, max_delta_e76, consistency, foreground,
                         best_contrast_ratio, relative_luminance
  · README             : version -> 1.1.0-corrected + a changelog note

(Upstream v1.1.0 already carries the 2 Luminance #FFFFFF + 3 individual black fixes and adds the NEOART 6901 (NART) series; this build adds the full SUP/NC2 series re-sample + the 2025-catalogue lightfastness correction on top.)

Note: re-saving via openpyxl simplifies workbook-level extras (conditional formatting,
the Dashboard chart may not survive). The DATA is authoritative; visuals are secondary.

Usage:  python3 build_corrected_xlsx.py
Output: ../reference/Caran_dAche_Master_Color_Index_v1.1.0-corrected.xlsx
"""
import json
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "Caran_dAche_Master_Color_Index_v1.1.0.xlsx")
RESAMPLED = os.path.join(HERE, "resampled_hex.json")
OUTDIR = os.path.abspath(os.path.join(HERE, "..", "reference"))
OUT = os.path.join(OUTDIR, "Caran_dAche_Master_Color_Index_v1.1.0-corrected.xlsx")
SERIES_ORDER = ["LUM", "PAB", "SUP", "MUS", "NC2", "NEO", "PSTP", "PSTC", "NART"]


def rgb_of(h):
    h = h.lstrip("#")
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def rgb_to_hsl(r, g, b):
    r, g, b = r / 255, g / 255, b / 255
    mx, mn = max(r, g, b), min(r, g, b)
    l = (mx + mn) / 2
    if mx == mn:
        return 0.0, 0.0, l * 100
    d = mx - mn
    s = d / (2 - mx - mn) if l > 0.5 else d / (mx + mn)
    if mx == r:
        h = (g - b) / d + (6 if g < b else 0)
    elif mx == g:
        h = (b - r) / d + 2
    else:
        h = (r - g) / d + 4
    return round(h * 60, 2), round(s * 100, 2), round(l * 100, 2)


def _lin(c):
    c /= 255.0
    return c / 12.92 if c <= 0.04045 else ((c + 0.055) / 1.055) ** 2.4


def rel_lum(r, g, b):
    return 0.2126 * _lin(r) + 0.7152 * _lin(g) + 0.0722 * _lin(b)


def lab_of(r, g, b):
    R, G, B = _lin(r), _lin(g), _lin(b)
    X = (R * 0.4124 + G * 0.3576 + B * 0.1805) / 0.95047
    Y = R * 0.2126 + G * 0.7152 + B * 0.0722
    Z = (R * 0.0193 + G * 0.1192 + B * 0.9505) / 1.08883

    def f(t):
        return t ** (1 / 3.0) if t > 0.008856 else 7.787 * t + 16 / 116.0
    fx, fy, fz = f(X), f(Y), f(Z)
    return (round(116 * fy - 16, 2), round(500 * (fx - fy), 2), round(200 * (fy - fz), 2))


def de76(la, lb):
    return sum((a - b) ** 2 for a, b in zip(la, lb)) ** 0.5


def contrasts(r, g, b):
    L = rel_lum(r, g, b)
    cb = (L + 0.05) / 0.05
    cw = 1.05 / (L + 0.05)
    fg = ("#000000", "Black") if cb >= cw else ("#FFFFFF", "White")
    return round(cb, 2), round(cw, 2), round(max(cb, cw), 2), round(L, 6), fg


def consistency(max_de):
    return "High" if max_de < 10 else "Medium" if max_de < 25 else "Low"


def hdr_idx(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value is not None}


def main():
    with open(RESAMPLED, encoding="utf-8") as f:
        rsp = json.load(f)["hex"]
    corr = {}  # (sid, code) -> hex (uppercase, #RRGGBB)
    for sid, m in rsp.items():
        for code, h in m.items():
            corr[(sid, str(code))] = "#" + h.lstrip("#").upper()

    wb = openpyxl.load_workbook(SRC, data_only=False)
    changed = {}

    # ---- Series_Color_Index ------------------------------------------------
    ws = wb["Series_Color_Index"]
    H = hdr_idx(ws)
    n = 0
    for row in ws.iter_rows(min_row=2):
        sid = row[H["series_id"] - 1].value
        code = str(row[H["color_code"] - 1].value)
        key = (sid, code)
        if key not in corr:
            continue
        hexv = corr[key]
        r, g, b = rgb_of(hexv)
        hh, ss, ll = rgb_to_hsl(r, g, b)
        L, a, bb = lab_of(r, g, b)
        cb, cw, best, lum, (fghex, fgname) = contrasts(r, g, b)

        def put(col, val):
            row[H[col] - 1].value = val
        put("css_hex_approx", hexv)
        put("rgb_r", r); put("rgb_g", g); put("rgb_b", b)
        put("hsl_h", hh); put("hsl_s", ss); put("hsl_l", ll)
        put("lab_l", L); put("lab_a", a); put("lab_b", bb)
        put("recommended_foreground_hex", fghex)
        put("recommended_foreground_name", fgname)
        put("best_contrast_ratio", best)
        put("contrast_with_black", cb); put("contrast_with_white", cw)
        put("relative_luminance", lum)
        put("wcag_aa_normal_text", "PASS" if best >= 4.5 else "FAIL")
        if "data_quality_status" in H:
            put("data_quality_status", "Corrected (SUP/NC2 re-sampled from official chart)")
        n += 1
    changed["Series_Color_Index"] = n

    # ---- Cross_Series_Map --------------------------------------------------
    ws = wb["Cross_Series_Map"]
    H = hdr_idx(ws)
    cross = {}  # code -> {sid: hex}
    n = 0
    for row in ws.iter_rows(min_row=2):
        code = str(row[H["color_code"] - 1].value)
        per = {}
        for sid in SERIES_ORDER:
            if sid not in H:
                continue
            cell = row[H[sid] - 1]
            val = cell.value
            if (sid, code) in corr and val not in (None, "—", "-", ""):
                cell.value = corr[(sid, code)]
                n += 1
            v = cell.value
            if v not in (None, "—", "-", ""):
                per[sid] = "#" + str(v).lstrip("#").upper()
        cross[code] = per
    changed["Cross_Series_Map"] = n

    # ---- Swatch_Reference --------------------------------------------------
    ws = wb["Swatch_Reference"]
    H = hdr_idx(ws)
    n = 0
    for row in ws.iter_rows(min_row=2):
        sid = row[H["series_id"] - 1].value
        code = str(row[H["color_code"] - 1].value)
        if (sid, code) not in corr:
            continue
        hexv = corr[(sid, code)]
        r, g, b = rgb_of(hexv)
        _, _, _, _, (fghex, _) = contrasts(r, g, b)
        row[H["css_hex_approx"] - 1].value = hexv
        row[H["recommended_foreground_hex"] - 1].value = fghex
        n += 1
    changed["Swatch_Reference"] = n

    # ---- Color_Master (recompute affected canonical rows) ------------------
    ws = wb["Color_Master"]
    H = hdr_idx(ws)
    affected_codes = {code for (sid, code) in corr}
    n = 0
    for row in ws.iter_rows(min_row=2):
        code = str(row[H["color_code"] - 1].value)
        if code not in affected_codes:
            continue
        per = cross.get(code, {})
        if not per:
            continue
        rgbs = [rgb_of(h) for h in per.values()]
        ra = round(sum(x[0] for x in rgbs) / len(rgbs))
        ga = round(sum(x[1] for x in rgbs) / len(rgbs))
        ba = round(sum(x[2] for x in rgbs) / len(rgbs))
        avg_hex = "#%02X%02X%02X" % (ra, ga, ba)
        labs = [lab_of(*x) for x in rgbs]
        max_de = max((de76(a, b) for i, a in enumerate(labs) for b in labs[i + 1:]), default=0.0)
        La, aa, ba2 = lab_of(ra, ga, ba)
        cb, cw, best, lum, (fghex, _) = contrasts(ra, ga, ba)

        def put(col, val):
            if col in H:
                row[H[col] - 1].value = val
        put("digital_reference_hex_avg", avg_hex)
        put("rgb_r_avg", ra); put("rgb_g_avg", ga); put("rgb_b_avg", ba)
        put("lab_l_avg", La); put("lab_a_avg", aa); put("lab_b_avg", ba2)
        put("max_delta_e76_between_series", round(max_de, 2))
        put("cross_series_consistency", consistency(max_de))
        put("recommended_foreground_hex", fghex)
        put("best_contrast_ratio", best)
        put("relative_luminance", lum)
        put("notes", "Recomputed after SUP/NC2 correction")
        n += 1
    changed["Color_Master"] = n

    # ---- Lightfastness (from the official 2025 Beaux-Arts catalogue) --------
    # The master index's lightfastness_rating was systematically ~2 stars too low
    # (PAB/SUP/NC2 collapsed to the minimum). See DESIGN.md §4.2. LUM untouched.
    lfp = os.path.join(HERE, "catalogue_lightfastness.json")
    LF = {}
    if os.path.exists(lfp):
        with open(lfp, encoding="utf-8") as f:
            for sid, m in json.load(f)["stars"].items():
                for code, st in m.items():
                    LF[(sid, str(code))] = st
    ws = wb["Series_Color_Index"]
    H = hdr_idx(ws)
    n = 0
    for row in ws.iter_rows(min_row=2):
        sid = row[H["series_id"] - 1].value
        code = str(row[H["color_code"] - 1].value)
        st = LF.get((sid, code))
        if st is None:
            continue
        smax = row[H["lightfastness_scale_max"] - 1].value or (5 if sid in ("MUS", "PSTP", "PSTC") else 3)
        row[H["lightfastness_rating"] - 1].value = "H" * st
        if "lightfastness_normalized_5" in H:
            row[H["lightfastness_normalized_5"] - 1].value = round(st / smax * 5, 2)
        n += 1
    changed["Lightfastness"] = n

    # ---- README (version + changelog) --------------------------------------
    # Base is upstream v1.1.0 (Luminance/black fixes + the new NEOART 6901 series). This build
    # layers the FULL SUP/NC2 re-sample + the 2025-catalogue lightfastness fix on top →
    # "1.1.0-corrected". Keep the upstream changelog rows; append ours.
    ws = wb["README"]
    if ws["A1"].value:
        ws["A1"].value = "Caran d’Ache Master Color Index v1.1.0-corrected"
    for row in ws.iter_rows(min_row=1, max_col=2):
        k = row[0].value
        if k == "Workbook version":
            row[1].value = "1.1.0-corrected"
        elif k == "Release date":
            row[1].value = "2026-07-22"
    last = ws.max_row + 1
    notes = [
        ("Revision", "v1.1.0-corrected (unofficial — app-author reference build on top of upstream v1.1.0)"),
        ("Correction", "Supracolor (SUP, 120) and Neocolor II (NC2, 84) were recorded systematically "
                       "too pale across the WHOLE palette (upstream v1.0.2 only fixed 3 individual black "
                       "swatches). Both series were re-sampled from the official Caran d'Ache colour charts "
                       "(median RGB of the swatch core, the method the workbook claims; validated: PAB/NEO/"
                       "PSTP matched at deltaE76 4-7, SUP/NC2 were off by 31-37)."),
        ("Corrected scope", "204 series-colours re-sampled; Cross_Series_Map, Swatch_Reference and all "
                            "SUP/NC2-bearing Color_Master rows recomputed for consistency (hex<->rgb verified)."),
        ("Lightfastness fix", "lightfastness_rating (and _normalized_5) were systematically ~2 stars too "
                              "low across MUS/PAB/SUP/NC2/NEO/PSTP/PSTC (PAB/SUP/NC2 collapsed every colour "
                              "to the minimum). Replaced with the official per-colour star ratings from the "
                              "Caran d'Ache Beaux-Arts 2025 catalogue. LUM (LFI/LFII) unchanged."),
        ("Build note", "Re-saved via openpyxl from v1.1.0; conditional formatting / dashboard visuals may "
                       "be simplified — the tabular DATA is authoritative. hex remains a screen "
                       "approximation, not an official Caran d'Ache RGB spec."),
    ]
    for i, (k, v) in enumerate(notes):
        ws.cell(row=last + i, column=1, value=k)
        ws.cell(row=last + i, column=2, value=v)

    os.makedirs(OUTDIR, exist_ok=True)
    wb.save(OUT)
    print("wrote", OUT)
    for s, c in changed.items():
        print(f"  {s}: {c} cells/rows corrected")


if __name__ == "__main__":
    main()
