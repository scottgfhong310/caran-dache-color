#!/usr/bin/env python3
"""
generate.py — 由 Caran_dAche_Master_Color_Index_v1.0.xlsx 產生前端靜態 registry。
==============================================================================
單一真相 = 這支 xlsx（放同資料夾）。改資料只需重跑本產生器，前端三支 cda-*.js
隨之更新，兩處不會分歧（比照 faber-castell-color 的 data/source/generate.js）。

輸出（寫進 ../../public/apps/caran-dache-color/data/）：
  cda-series.js    window.CDA_SERIES   — 8 產品系列 registry
  cda-colors.js    window.CDA_COLORS   — 764 系列色列（每個顏色「在某系列裡」一列）
                   window.CDA_META     — 版本 / 來源 / 統計 / 系列順序
  cda-canonical.js window.CDA_CANONICAL — 227 正典色碼 ＋ 同碼跨系列 hex 對照

執行： python3 generate.py       （需 openpyxl；僅建置期用，不進前端）

資料語言：前端只取 en（canonical）＋ zh-Hant ＋ ja 三語（對齊家族三語慣例）；
xlsx 另有 fr/de/it/es/pt/ko，需要時再擴充。
hex 一律正規化為小寫 #rrggbb。空字串 / '—' 一律轉 None（省略該鍵）。
"""

import json
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Caran_dAche_Master_Color_Index_v1.0.2.xlsx")
OUT = os.path.abspath(os.path.join(HERE, "..", "..", "public", "apps", "caran-dache-color", "data"))

SERIES_ORDER = ["LUM", "PAB", "SUP", "MUS", "NC2", "NEO", "PSTP", "PSTC"]

# --- Corrections applied on top of the source master index (documented; DESIGN.md §4.1) ----
# The SUP (Supracolor) and NC2 (Neocolor II) series are systematically mis-extracted in the
# master index — recorded far too pale/washed-out (verified vs the official chart PDFs: with
# the same pixel-sampling method, PAB/NEO/PSTP matched at ΔE76 4–7, but SUP/NC2 were off by
# ~31–37). extract_charts.py re-samples both whole series from the official Supracolor /
# Neocolor II charts; resampled_hex.json holds the corrected per-code hex. (The earlier two
# Luminance #FFFFFF swatches — 009 Black, 639 Dark indigo — were fixed upstream in v1.0.1,
# so no manual Luminance override is needed here.)
_RESAMPLED = {}
_NOTE = {
    "SUP": "corrected: the Supracolor series was recorded too pale in the source; "
           "re-sampled from the official Supracolor colour chart",
    "NC2": "corrected: the Neocolor II series was recorded too pale in the source; "
           "re-sampled from the official Neocolor II colour chart",
}
_rp = os.path.join(HERE, "resampled_hex.json")
if os.path.exists(_rp):
    with open(_rp, encoding="utf-8") as _f:
        for _sid, _m in json.load(_f).get("hex", {}).items():
            for _code, _hex in _m.items():
                _RESAMPLED[(_sid, _code)] = (_hex, _NOTE.get(_sid, "corrected: re-sampled from official chart"))


def override_for(sid, code):
    return _RESAMPLED.get((sid, code))


_OVERRIDDEN_CODES = {code for (_sid, code) in _RESAMPLED}


def _rgb_of(hexv):
    return int(hexv[1:3], 16), int(hexv[3:5], 16), int(hexv[5:7], 16)


def _lab(r, g, b):
    def lin(c):
        c /= 255.0
        return c / 12.92 if c <= 0.04045 else ((c + 0.055) / 1.055) ** 2.4
    R, G, B = lin(r), lin(g), lin(b)
    X = (R * 0.4124 + G * 0.3576 + B * 0.1805) / 0.95047
    Y = (R * 0.2126 + G * 0.7152 + B * 0.0722)
    Z = (R * 0.0193 + G * 0.1192 + B * 0.9505) / 1.08883

    def f(t):
        return t ** (1 / 3.0) if t > 0.008856 else (7.787 * t + 16 / 116.0)
    fx, fy, fz = f(X), f(Y), f(Z)
    return (116 * fy - 16, 500 * (fx - fy), 200 * (fy - fz))


def _de76(h1, h2):
    l1 = _lab(*_rgb_of(h1))
    l2 = _lab(*_rgb_of(h2))
    return sum((a - b) ** 2 for a, b in zip(l1, l2)) ** 0.5


def _consistency(max_de):
    return "High" if max_de < 10 else "Medium" if max_de < 25 else "Low"


def s(v):
    """Trim to clean string, or None for empty / dash placeholders."""
    if v is None:
        return None
    t = str(v).strip()
    if t == "" or t == "—" or t == "-":
        return None
    return t


def num(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 4)
    except (ValueError, TypeError):
        return None


def hexnorm(v):
    t = s(v)
    if not t:
        return None
    m = re.match(r"^#?([0-9a-fA-F]{6})$", t)
    return "#" + m.group(1).lower() if m else None


def sheet_rows(wb, name):
    ws = wb[name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    return idx, rows[1:]


def col(idx, row, name):
    i = idx.get(name)
    return row[i] if i is not None and i < len(row) else None


def js_header(*lines):
    return "/* " + "\n * ".join(lines) + "\n */\n"


def dump(var, data):
    return var + " = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # ---- README meta -------------------------------------------------------
    meta = {}
    for r in wb["README"].iter_rows(values_only=True):
        if r and len(r) >= 2 and r[0] and r[1] is not None:
            meta[str(r[0]).strip()] = str(r[1]).strip()

    # ---- Series_Registry → CDA_SERIES -------------------------------------
    sidx, srows = sheet_rows(wb, "Series_Registry")
    series = []
    for r in srows:
        sid = s(col(sidx, r, "series_id"))
        if not sid:
            continue
        series.append({
            "id": sid,
            "name": s(col(sidx, r, "series_name")),
            "medium": s(col(sidx, r, "medium_type")),
            "grade": s(col(sidx, r, "product_grade")),
            "count": num(col(sidx, r, "official_colour_count")),
            "lfStandard": s(col(sidx, r, "lightfastness_standard")),
            "ratingScale": s(col(sidx, r, "rating_scale")),
            "sharedWith": s(col(sidx, r, "shared_palette_with")),
        })
    series.sort(key=lambda x: SERIES_ORDER.index(x["id"]) if x["id"] in SERIES_ORDER else 99)

    # ---- Series_Color_Index → CDA_COLORS ----------------------------------
    cidx, crows = sheet_rows(wb, "Series_Color_Index")
    colors = []
    for r in crows:
        sid = s(col(cidx, r, "series_id"))
        code = s(col(cidx, r, "color_code"))
        hexv = hexnorm(col(cidx, r, "css_hex_approx"))
        if not (sid and code and hexv):
            continue
        override = override_for(sid, code)
        note = None
        if override:
            hexv, note = override[0], override[1]
        R, G, B = _rgb_of(hexv)
        c = {
            "id": s(col(cidx, r, "series_color_id")),
            "seriesId": sid,
            "code": code,
            "order": num(col(cidx, r, "palette_order")),
            "name": s(col(cidx, r, "color_name_en")),
            "nameZh": s(col(cidx, r, "color_name_zh_tw")),
            "nameJa": s(col(cidx, r, "color_name_ja")),
            "hex": hexv,
            "r": R, "g": G, "b": B,
            "lf": s(col(cidx, r, "lightfastness_rating")),
            "lfNorm": num(col(cidx, r, "lightfastness_normalized_5")),
            "lfMax": num(col(cidx, r, "lightfastness_scale_max")),
            "lfStd": s(col(cidx, r, "lightfastness_standard")),
            "pig": s(col(cidx, r, "pigment_index")),
            "pigN": num(col(cidx, r, "pigment_count")),
            "wcag": s(col(cidx, r, "wcag_aa_normal_text")),
            "contrast": num(col(cidx, r, "best_contrast_ratio")),
            "canon": s(col(cidx, r, "canonical_code_id")),
            "cssVar": s(col(cidx, r, "css_variable")),
            "note": note,
        }
        colors.append({k: v for k, v in c.items() if v is not None})

    # ---- Color_Master + Cross_Series_Map → CDA_CANONICAL ------------------
    xidx, xrows = sheet_rows(wb, "Cross_Series_Map")
    cross = {}
    for r in xrows:
        code = s(col(xidx, r, "color_code"))
        if not code:
            continue
        per = {}
        for sid in SERIES_ORDER:
            h = hexnorm(col(xidx, r, sid))
            override = override_for(sid, code)
            if override:
                h = override[0]
            if h:
                per[sid] = h
        cross[code] = per

    midx, mrows = sheet_rows(wb, "Color_Master")
    canonical = []
    for r in mrows:
        code = s(col(midx, r, "color_code"))
        if not code:
            continue
        slist = s(col(midx, r, "series_list"))
        c = {
            "code": code,
            "name": s(col(midx, r, "canonical_name_en")),
            "nameZh": s(col(midx, r, "canonical_name_zh_tw")),
            "nameJa": s(col(midx, r, "canonical_name_ja")),
            "seriesCount": num(col(midx, r, "series_count")),
            "seriesList": [p.strip() for p in re.split(r"\|", slist)] if slist else None,
            "pigments": s(col(midx, r, "pigment_variants")),
            "avgHex": hexnorm(col(midx, r, "digital_reference_hex_avg")),
            "maxDeltaE76": num(col(midx, r, "max_delta_e76_between_series")),
            "consistency": s(col(midx, r, "cross_series_consistency")),
            "series": cross.get(code) or None,
        }
        # Recompute the affected canonical rows from the corrected per-series hexes so the
        # cross-series average / ΔE76 / consistency reflect the corrections (the source values
        # here were polluted by the mis-extracted SUP/NC2 pale hexes).
        if code in _OVERRIDDEN_CODES and c["series"]:
            hexes = list(c["series"].values())
            rs = [_rgb_of(h) for h in hexes]
            ra = round(sum(x[0] for x in rs) / len(rs))
            ga = round(sum(x[1] for x in rs) / len(rs))
            ba = round(sum(x[2] for x in rs) / len(rs))
            c["avgHex"] = "#%02x%02x%02x" % (ra, ga, ba)
            max_de = max((_de76(a, b) for i, a in enumerate(hexes) for b in hexes[i + 1:]),
                         default=0.0)
            c["maxDeltaE76"] = round(max_de, 2)
            c["consistency"] = _consistency(max_de)
            c["note"] = ("cross-series average recomputed after correcting mis-extracted "
                         "series values in the source (see series-colour note)")
        canonical.append({k: v for k, v in c.items() if v is not None})

    # ---- META --------------------------------------------------------------
    cda_meta = {
        "source": os.path.basename(XLSX),
        "version": meta.get("Workbook version", "1.0"),
        "releaseDate": meta.get("Release date", ""),
        "brand": "Caran d’Ache",
        "seriesOrder": SERIES_ORDER,
        "total": len(colors),
        "canonicalTotal": len(canonical),
        "langs": ["en", "zh-Hant", "ja"],
        "note": "Median RGB sampled from official PDF swatches; screen-reference approximation only, not an official RGB/HEX specification.",
    }

    os.makedirs(OUT, exist_ok=True)

    with open(os.path.join(OUT, "cda-series.js"), "w", encoding="utf-8") as f:
        f.write(js_header(
            "Caran d’Ache product series registry — generated from",
            "data/source/" + os.path.basename(XLSX) + " (do not hand-edit).",
            "8 series (LUM/PAB/SUP/MUS/NC2/NEO/PSTP/PSTC); see data/source/generate.py.",
        ))
        f.write(dump("window.CDA_SERIES", series))

    with open(os.path.join(OUT, "cda-colors.js"), "w", encoding="utf-8") as f:
        f.write(js_header(
            "Caran d’Ache series colours — generated (do not hand-edit).",
            "Source: Caran_dAche_Master_Color_Index_v1.0.xlsx (Series_Color_Index).",
            "One row per colour-in-a-series (" + str(len(colors)) + " rows across 8 series).",
            "Fields: id, seriesId, code, order, name/nameZh/nameJa, hex, r/g/b,",
            "  lf/lfNorm/lfMax/lfStd, pig/pigN, wcag, contrast, canon, cssVar.",
            "hex = median RGB sampled from official PDF swatches, approximate (not official).",
        ))
        f.write(dump("window.CDA_META", cda_meta))
        f.write(dump("window.CDA_COLORS", colors))

    with open(os.path.join(OUT, "cda-canonical.js"), "w", encoding="utf-8") as f:
        f.write(js_header(
            "Caran d’Ache canonical colour codes — generated (do not hand-edit).",
            "Source: Color_Master + Cross_Series_Map (" + str(len(canonical)) + " codes).",
            "Each code de-duplicates a shared colour number across series; `series` maps",
            "  series_id -> that series' sampled hex for the same code (cross-series view).",
            "Fields: code, name/nameZh/nameJa, seriesCount, seriesList, pigments,",
            "  avgHex, maxDeltaE76, consistency, series{ SID: hex }.",
        ))
        f.write(dump("window.CDA_CANONICAL", canonical))

    print(f"series:    {len(series)}")
    print(f"colors:    {len(colors)}")
    print(f"canonical: {len(canonical)}")
    for fn in ("cda-series.js", "cda-colors.js", "cda-canonical.js"):
        p = os.path.join(OUT, fn)
        print(f"  {fn:18} {os.path.getsize(p):>8,} bytes")


if __name__ == "__main__":
    main()
